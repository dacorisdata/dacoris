from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Type, TypeVar

from openpyxl import load_workbook

from .schemas import (
    FMSFeeStructure,
    FMSStudentAccount,
    FMSTransaction,
    HRStaff,
    HRSupervisorAssignment,
    JourneyStageStatus,
    LMSCourse,
    LMSEnrolment,
    LMSJourney,
    LMSProgramme,
    LMSProgrammeRule,
    SISCohort,
    SISStudent,
)

T = TypeVar("T")

DEFAULT_EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "DACORIS_IS_v2.xlsx"

JOURNEY_STAGE_NAMES = [
    "Admission & Enrolment",
    "Coursework",
    "Supervisor Assignment",
    "Proposal Development",
    "Proposal Defense",
    "Data Collection",
    "Thesis Writing",
    "Publication",
    "Thesis Defense",
    "Graduation",
]

_cache: Dict[str, Any] = {"mtime": None, "repo": None}


def _parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _parse_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _yes_no(value: Any) -> bool:
    text = _cell_str(value).lower()
    return text in ("yes", "y", "true", "1")


def _headers(row) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is not None and str(cell).strip():
            mapping[str(cell).strip()] = idx
    return mapping


def _row_values(row, headers: Dict[str, int], *keys: str, default: Any = "") -> List[Any]:
    return [row[headers[k]] if k in headers and headers[k] < len(row) else default for k in keys]


def _parse_stage_header_name(raw: Any) -> str:
    text = _cell_str(raw).replace("\n", " ").strip()
    if not text:
        return ""
    return text


def _journey_stage_layout(header_row) -> List[tuple[int, int, str, str]]:
    """Return (start_col, width, extra_kind, stage_name) for each stage column group."""
    summary_start = next(
        (idx for idx, cell in enumerate(header_row) if _cell_str(cell) == "Current Stage"),
        len(header_row),
    )
    layout: List[tuple[int, int, str, str]] = []
    col = 6
    while col < summary_start:
        next_header = _cell_str(header_row[col + 1]) if col + 1 < len(header_row) else ""
        if next_header == "Units Done":
            width, extra_kind = 3, "units"
        elif next_header == "Pub Count":
            width, extra_kind = 3, "pub"
        else:
            width, extra_kind = 2, "date"
        layout.append((col, width, extra_kind, ""))
        col += width
    return layout


class ExcelISRepository:
    def __init__(self, excel_path: Optional[Path] = None):
        self.excel_path = Path(excel_path or os.getenv("DACORIS_IS_EXCEL_PATH", DEFAULT_EXCEL_PATH))
        self._students: List[SISStudent] = []
        self._cohorts: List[SISCohort] = []
        self._programmes: List[LMSProgramme] = []
        self._courses: List[LMSCourse] = []
        self._enrolments: List[LMSEnrolment] = []
        self._rules: List[LMSProgrammeRule] = []
        self._journeys: List[LMSJourney] = []
        self._fee_structures: List[FMSFeeStructure] = []
        self._accounts: List[FMSStudentAccount] = []
        self._transactions: List[FMSTransaction] = []
        self._staff: List[HRStaff] = []
        self._assignments: List[HRSupervisorAssignment] = []
        self._load()

    def _load(self) -> None:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"DACORIS IS Excel file not found: {self.excel_path}")

        wb = load_workbook(self.excel_path, read_only=True, data_only=True)
        self._students = self._load_students(wb["SIS_Students"])
        self._cohorts = self._load_cohorts(wb["SIS_Cohorts"])
        self._programmes = self._load_programmes(wb["LMS_Programmes"])
        self._courses = self._load_courses(wb["LMS_Courses"])
        self._enrolments = self._load_enrolments(wb["LMS_Enrolments"])
        self._rules = self._load_rules(wb["LMS_Programme_Rules"])
        self._journeys = self._load_journeys(wb["LMS_Journey"])
        self._fee_structures = self._load_fee_structures(wb["FMS_Fee_Structure"])
        self._accounts = self._load_accounts(wb["FMS_Student_Accounts"])
        self._transactions = self._load_transactions(wb["FMS_Transactions"])
        self._staff = self._load_staff(wb["HR_Staff"])
        self._assignments = self._load_assignments(wb["HR_Supervisor_Assignments"])
        wb.close()

    def _load_students(self, ws) -> List[SISStudent]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        students: List[SISStudent] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Student ID", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            students.append(
                SISStudent(
                    student_id=_cell_str(v("Student ID")[0]),
                    first_name=_cell_str(v("First Name")[0]),
                    last_name=_cell_str(v("Last Name")[0]),
                    full_name=_cell_str(v("Full Name")[0]),
                    email=_cell_str(v("Email")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    domain=_cell_str(v("Domain")[0]),
                    gender=_cell_str(v("Gender")[0]),
                    nationality=_cell_str(v("Nationality")[0]),
                    student_type=_cell_str(v("Student Type")[0]),
                    degree_level=_cell_str(v("Degree Level")[0]),
                    programme_code=_cell_str(v("Programme Code")[0]),
                    programme_name=_cell_str(v("Programme Name")[0]),
                    department=_cell_str(v("Department")[0]),
                    cohort_year=_parse_int(v("Cohort Year")[0]),
                    enrolment_date=_parse_date(v("Enrolment Date")[0]),
                    study_mode=_cell_str(v("Study Mode")[0]),
                    expected_graduation_date=_parse_date(v("Expected Graduation Date")[0]),
                    current_stage_no=_parse_int(v("Current Stage No.")[0]),
                    current_stage_name=_cell_str(v("Current Stage Name")[0]),
                    orcid_placeholder=_cell_str(v("ORCID Placeholder")[0]),
                    status=_cell_str(v("Status")[0]),
                )
            )
        return students

    def _load_cohorts(self, ws) -> List[SISCohort]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[SISCohort] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Cohort ID", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                SISCohort(
                    cohort_id=_cell_str(v("Cohort ID")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    cohort_year=_parse_int(v("Cohort Year")[0]),
                    degree_level=_cell_str(v("Degree Level")[0]),
                    programme_code=_cell_str(v("Programme Code")[0]),
                    programme_name=_cell_str(v("Programme Name")[0]),
                    start_date=_parse_date(v("Start Date")[0]),
                    expected_end_date=_parse_date(v("Expected End Date")[0]),
                    total_enrolled=_parse_int(v("Total Enrolled")[0]),
                    msc_count=_parse_int(v("MSc Count")[0]),
                    phd_count=_parse_int(v("PhD Count")[0]),
                    local_count=_parse_int(v("Local")[0]),
                    international_count=_parse_int(v("International")[0]),
                    active=_parse_int(v("Active")[0]),
                    status=_cell_str(v("Status")[0]),
                )
            )
        return items

    def _load_programmes(self, ws) -> List[LMSProgramme]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[LMSProgramme] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Programme Code", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                LMSProgramme(
                    programme_code=_cell_str(v("Programme Code")[0]),
                    programme_name=_cell_str(v("Programme Name")[0]),
                    degree_level=_cell_str(v("Degree Level")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    department=_cell_str(v("Department")[0]),
                    duration_years=_parse_float(v("Duration (Years)")[0]),
                    total_credits=_parse_int(v("Total Credits")[0]),
                    pub_requirement=_parse_int(v("Pub. Requirement")[0]),
                    min_coursework_units=_parse_int(v("Min Coursework Units")[0]),
                    supervisor_required=_yes_no(v("Supervisor Required")[0]),
                    ethics_required=_yes_no(v("Ethics Required")[0]),
                    dmp_required=_yes_no(v("DMP Required")[0]),
                    thesis_required=_yes_no(v("Thesis Required")[0]),
                    graduation_gate=_yes_no(v("Graduation Gate")[0]),
                    status=_cell_str(v("Status")[0]),
                )
            )
        return items

    def _load_courses(self, ws) -> List[LMSCourse]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[LMSCourse] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Course Code", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                LMSCourse(
                    course_code=_cell_str(v("Course Code")[0]),
                    course_name=_cell_str(v("Course Name")[0]),
                    credits=_parse_int(v("Credits")[0]),
                    programme_code=_cell_str(v("Programme Code")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    year_of_study=_parse_int(v("Year of Study")[0]),
                    semester=_parse_int(v("Semester")[0]),
                    course_type=_cell_str(v("Type")[0]),
                    status=_cell_str(v("Status")[0]),
                )
            )
        return items

    def _load_enrolments(self, ws) -> List[LMSEnrolment]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[LMSEnrolment] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Enrolment ID", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                LMSEnrolment(
                    enrolment_id=_cell_str(v("Enrolment ID")[0]),
                    student_id=_cell_str(v("Student ID")[0]),
                    full_name=_cell_str(v("Full Name")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    programme_code=_cell_str(v("Programme Code")[0]),
                    course_code=_cell_str(v("Course Code")[0]),
                    course_name=_cell_str(v("Course Name")[0]),
                    academic_year=_cell_str(v("Academic Year")[0]),
                    semester=_parse_int(v("Semester")[0]),
                    mark_pct=_parse_float(v("Mark (%)")[0]),
                    grade=_cell_str(v("Grade")[0]),
                    status=_cell_str(v("Status")[0]),
                )
            )
        return items

    def _load_rules(self, ws) -> List[LMSProgrammeRule]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[LMSProgrammeRule] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Rule ID", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                LMSProgrammeRule(
                    rule_id=_cell_str(v("Rule ID")[0]),
                    programme_code=_cell_str(v("Programme Code")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    degree_level=_cell_str(v("Degree Level")[0]),
                    rule_type=_cell_str(v("Rule Type")[0]),
                    rule_description=_cell_str(v("Rule Description")[0]),
                    min_value=_parse_float(v("Min Value")[0]),
                    acceptable_evidence=_cell_str(v("Acceptable Evidence")[0]),
                    mandatory=_yes_no(v("Mandatory")[0]) if v("Mandatory")[0] not in (None, "") else True,
                    managed_by=_cell_str(v("Managed By")[0]),
                )
            )
        return items

    def _load_journeys(self, ws) -> List[LMSJourney]:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return []

        title_row = rows[0]
        header_row = rows[1]
        summary_headers = _headers(header_row)
        stage_layout = _journey_stage_layout(header_row)

        for idx, (start_col, _width, _extra_kind, _name) in enumerate(stage_layout):
            if start_col < len(title_row):
                stage_layout[idx] = (
                    start_col,
                    _width,
                    _extra_kind,
                    _parse_stage_header_name(title_row[start_col]),
                )

        def summary_value(row, key: str, default: Any = ""):
            col = summary_headers.get(key)
            if col is None or col >= len(row):
                return default
            return row[col]

        items: List[LMSJourney] = []
        for row in rows[2:]:
            if not row or not row[0]:
                continue
            stages: List[JourneyStageStatus] = []
            for stage_no, (start_col, width, extra_kind, stage_title) in enumerate(stage_layout, start=1):
                status = _cell_str(row[start_col]) if start_col < len(row) else ""
                stage_date = None
                extra: Dict[str, Any] = {}
                if extra_kind == "units" and start_col + 1 < len(row):
                    extra["units_done"] = _cell_str(row[start_col + 1])
                    stage_date = _parse_date(row[start_col + 2]) if start_col + 2 < len(row) else None
                elif extra_kind == "pub" and start_col + 1 < len(row):
                    extra["pub_count"] = _parse_int(row[start_col + 1])
                    stage_date = _parse_date(row[start_col + 2]) if start_col + 2 < len(row) else None
                elif width >= 2 and start_col + 1 < len(row):
                    stage_date = _parse_date(row[start_col + 1])
                stage_name = stage_title or (
                    JOURNEY_STAGE_NAMES[stage_no - 1] if stage_no <= len(JOURNEY_STAGE_NAMES) else f"Stage {stage_no}"
                )
                stages.append(
                    JourneyStageStatus(
                        stage_no=stage_no,
                        stage_name=stage_name,
                        status=status,
                        stage_date=stage_date,
                        extra=extra,
                    )
                )
            items.append(
                LMSJourney(
                    student_id=_cell_str(row[0]),
                    full_name=_cell_str(row[1]),
                    institution=_cell_str(row[2]),
                    programme=_cell_str(row[3]),
                    level=_cell_str(row[4]),
                    cohort=_parse_int(row[5]),
                    current_stage=_cell_str(summary_value(row, "Current Stage")),
                    overall_status=_cell_str(summary_value(row, "Overall Status")),
                    days_overdue=_parse_int(summary_value(row, "Days Overdue", None)),
                    risk_level=_cell_str(summary_value(row, "Risk Level")),
                    lead_supervisor=_cell_str(summary_value(row, "Lead Supervisor")),
                    notes=_cell_str(summary_value(row, "Notes")),
                    expected_graduation=_parse_date(summary_value(row, "Grad Date", None)),
                    stages=stages,
                )
            )
        return items

    def _load_fee_structures(self, ws) -> List[FMSFeeStructure]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[FMSFeeStructure] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Fee Code", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                FMSFeeStructure(
                    fee_code=_cell_str(v("Fee Code")[0]),
                    programme_code=_cell_str(v("Programme Code")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    degree_level=_cell_str(v("Degree Level")[0]),
                    academic_year=_cell_str(v("Academic Year")[0]),
                    tuition_kes=_parse_float(v("Tuition (KES)")[0]),
                    registration_kes=_parse_float(v("Registration (KES)")[0]),
                    thesis_exam_kes=_parse_float(v("Thesis/Examination (KES)")[0]),
                    library_kes=_parse_float(v("Library (KES)")[0]),
                    technology_levy_kes=_parse_float(v("Technology Levy (KES)")[0]),
                    total_annual_kes=_parse_float(v("Total Annual (KES)")[0]),
                    currency=_cell_str(v("Currency")[0]) or "KES",
                )
            )
        return items

    def _load_accounts(self, ws) -> List[FMSStudentAccount]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[FMSStudentAccount] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Account ID", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                FMSStudentAccount(
                    account_id=_cell_str(v("Account ID")[0]),
                    student_id=_cell_str(v("Student ID")[0]),
                    full_name=_cell_str(v("Full Name")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    programme=_cell_str(v("Programme")[0]),
                    degree_level=_cell_str(v("Degree Level")[0]),
                    cohort=_parse_int(v("Cohort")[0]),
                    expected_graduation_date=_parse_date(v("Expected Graduation Date")[0]),
                    annual_fee_kes=_parse_float(v("Annual Fee (KES)")[0]),
                    duration_yrs=_parse_int(v("Duration (Yrs)")[0]),
                    total_programme_fee_kes=_parse_float(v("Total Programme Fee (KES)")[0]),
                    scholarship_kes=_parse_float(v("Scholarship (KES)")[0]),
                    net_payable_kes=_parse_float(v("Net Payable (KES)")[0]),
                    amount_paid_kes=_parse_float(v("Amount Paid (KES)")[0]),
                    outstanding_kes=_parse_float(v("Outstanding (KES)")[0]),
                    payment_status=_cell_str(v("Payment Status")[0]),
                    last_payment_date=_parse_date(v("Last Payment Date")[0]),
                    finance_clearance=_yes_no(v("Finance Clearance")[0]),
                )
            )
        return items

    def _load_transactions(self, ws) -> List[FMSTransaction]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[FMSTransaction] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Transaction ID", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                FMSTransaction(
                    transaction_id=_cell_str(v("Transaction ID")[0]),
                    student_id=_cell_str(v("Student ID")[0]),
                    full_name=_cell_str(v("Full Name")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    transaction_date=_parse_date(v("Date")[0]),
                    amount_kes=_parse_float(v("Amount (KES)")[0]),
                    payment_method=_cell_str(v("Payment Method")[0]),
                    reference_no=_cell_str(v("Reference No.")[0]),
                    academic_year=_cell_str(v("Academic Year")[0]),
                    fee_type=_cell_str(v("Fee Type")[0]),
                    status=_cell_str(v("Status")[0]),
                    receipt_no=_cell_str(v("Receipt No.")[0]),
                )
            )
        return items

    def _load_staff(self, ws) -> List[HRStaff]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[HRStaff] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Staff ID", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                HRStaff(
                    staff_id=_cell_str(v("Staff ID")[0]),
                    first_name=_cell_str(v("First Name")[0]),
                    last_name=_cell_str(v("Last Name")[0]),
                    full_name=_cell_str(v("Full Name")[0]),
                    email=_cell_str(v("Email")[0]).lower(),
                    institution=_cell_str(v("Institution")[0]),
                    role=_cell_str(v("Role")[0]),
                    department=_cell_str(v("Department")[0]),
                    staff_type=_cell_str(v("Staff Type")[0]),
                    specialization=_cell_str(v("Specialization")[0]),
                    active_msc_students=_parse_int(v("Active MSc Students")[0]),
                    active_phd_students=_parse_int(v("Active PhD Students")[0]),
                    total_supervisees=_parse_int(v("Total Supervisees")[0]),
                    employment_type=_cell_str(v("Employment Type")[0]),
                    join_date=_parse_date(v("Join Date")[0]),
                    status=_cell_str(v("Status")[0]),
                )
            )
        return items

    def _load_assignments(self, ws) -> List[HRSupervisorAssignment]:
        rows = list(ws.iter_rows(values_only=True))
        headers = _headers(rows[0])
        items: List[HRSupervisorAssignment] = []
        for row in rows[1:]:
            if not row or not row[headers.get("Assignment ID", 0)]:
                continue
            v = lambda *keys: _row_values(row, headers, *keys)
            items.append(
                HRSupervisorAssignment(
                    assignment_id=_cell_str(v("Assignment ID")[0]),
                    student_id=_cell_str(v("Student ID")[0]),
                    student_name=_cell_str(v("Student Name")[0]),
                    institution=_cell_str(v("Institution")[0]),
                    programme=_cell_str(v("Programme")[0]),
                    level=_cell_str(v("Level")[0]),
                    lead_supervisor_id=_cell_str(v("Lead Supervisor ID")[0]),
                    lead_supervisor_name=_cell_str(v("Lead Supervisor Name")[0]),
                    lead_supervisor_email=_cell_str(v("Lead Supervisor Email")[0]).lower(),
                    co_supervisor_id=_cell_str(v("Co-Supervisor ID")[0]),
                    co_supervisor_name=_cell_str(v("Co-Supervisor Name")[0]),
                    co_supervisor_email=_cell_str(v("Co-Supervisor Email")[0]).lower(),
                    appointment_date=_parse_date(v("Appointment Date")[0]),
                    status=_cell_str(v("Status")[0]),
                    conflict_declared=_yes_no(v("Conflict Declared")[0]),
                    notes=_cell_str(v("Notes")[0]),
                )
            )
        return items

    def _filter_institution(self, institution_name: Optional[str], domain: Optional[str], items: List[T]) -> List[T]:
        if not institution_name and not domain:
            return items
        filtered: List[T] = []
        for item in items:
            inst = getattr(item, "institution", "") or ""
            dom = getattr(item, "domain", "") or ""
            if institution_name and inst.lower() == institution_name.lower():
                filtered.append(item)
            elif domain and dom.lower() == domain.lower():
                filtered.append(item)
            elif institution_name and institution_name.lower() in inst.lower():
                filtered.append(item)
        return filtered

    def get_students(self, institution_name: Optional[str] = None, domain: Optional[str] = None) -> List[SISStudent]:
        return self._filter_institution(institution_name, domain, self._students)

    def get_student(self, student_id: str, institution_name: Optional[str] = None, domain: Optional[str] = None) -> Optional[SISStudent]:
        for student in self.get_students(institution_name, domain):
            if student.student_id == student_id:
                return student
        return None

    def get_cohorts(self, institution_name: Optional[str] = None, domain: Optional[str] = None) -> List[SISCohort]:
        return self._filter_institution(institution_name, domain, self._cohorts)

    def get_programmes(self, institution_name: Optional[str] = None, domain: Optional[str] = None) -> List[LMSProgramme]:
        return self._filter_institution(institution_name, domain, self._programmes)

    def get_programme(self, programme_code: str, institution_name: Optional[str] = None, domain: Optional[str] = None) -> Optional[LMSProgramme]:
        for prog in self.get_programmes(institution_name, domain):
            if prog.programme_code == programme_code:
                return prog
        return None

    def get_enrolments(self, student_id: str, institution_name: Optional[str] = None, domain: Optional[str] = None) -> List[LMSEnrolment]:
        return [
            e for e in self._filter_institution(institution_name, domain, self._enrolments)
            if e.student_id == student_id
        ]

    def get_programme_rules(self, programme_code: str, institution_name: Optional[str] = None, domain: Optional[str] = None) -> List[LMSProgrammeRule]:
        return [
            r for r in self._filter_institution(institution_name, domain, self._rules)
            if r.programme_code == programme_code
        ]

    def get_journey(self, student_id: str, institution_name: Optional[str] = None, domain: Optional[str] = None) -> Optional[LMSJourney]:
        for journey in self._filter_institution(institution_name, domain, self._journeys):
            if journey.student_id == student_id:
                return journey
        return None

    def get_journeys(self, institution_name: Optional[str] = None, domain: Optional[str] = None) -> List[LMSJourney]:
        return self._filter_institution(institution_name, domain, self._journeys)

    def get_finance_account(self, student_id: str, institution_name: Optional[str] = None, domain: Optional[str] = None) -> Optional[FMSStudentAccount]:
        for account in self._filter_institution(institution_name, domain, self._accounts):
            if account.student_id == student_id:
                return account
        return None

    def get_transactions(self, student_id: str, institution_name: Optional[str] = None, domain: Optional[str] = None) -> List[FMSTransaction]:
        return [
            t for t in self._filter_institution(institution_name, domain, self._transactions)
            if t.student_id == student_id
        ]

    def get_staff(self, staff_id: Optional[str] = None, email: Optional[str] = None, institution_name: Optional[str] = None, domain: Optional[str] = None) -> Optional[HRStaff]:
        for member in self._filter_institution(institution_name, domain, self._staff):
            if staff_id and member.staff_id == staff_id:
                return member
            if email and member.email.lower() == email.lower():
                return member
        return None

    def get_staff_list(self, institution_name: Optional[str] = None, domain: Optional[str] = None) -> List[HRStaff]:
        return self._filter_institution(institution_name, domain, self._staff)

    def get_supervisor_assignments(
        self,
        student_id: Optional[str] = None,
        staff_id: Optional[str] = None,
        institution_name: Optional[str] = None,
        domain: Optional[str] = None,
    ) -> List[HRSupervisorAssignment]:
        items = self._filter_institution(institution_name, domain, self._assignments)
        if student_id:
            items = [a for a in items if a.student_id == student_id]
        if staff_id:
            items = [
                a for a in items
                if a.lead_supervisor_id == staff_id or a.co_supervisor_id == staff_id
            ]
        return items

    def get_staff_workload(self, staff_id: str, institution_name: Optional[str] = None, domain: Optional[str] = None) -> Optional[HRStaff]:
        return self.get_staff(staff_id=staff_id, institution_name=institution_name, domain=domain)


def get_excel_repository() -> ExcelISRepository:
    excel_path = Path(os.getenv("DACORIS_IS_EXCEL_PATH", DEFAULT_EXCEL_PATH))
    mtime = excel_path.stat().st_mtime if excel_path.exists() else None
    if _cache["repo"] is None or _cache["mtime"] != mtime:
        _cache["repo"] = ExcelISRepository(excel_path)
        _cache["mtime"] = mtime
    return _cache["repo"]
